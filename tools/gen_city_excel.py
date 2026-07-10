#!/usr/bin/env python3
"""Generate QWeather China city list Excel from local CSV cache."""
import csv, os, sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUTPUT = os.path.join(os.path.expanduser("~"), "Desktop", "和风天气中国城市码.xlsx")

HEADERS_ZH = [
    "城市ID", "英文名", "中文名", "ISO代码", "国家(英)", "国家(中)",
    "省份(英)", "省份(中)", "城市(英)", "城市(中)",
    "时区", "纬度", "经度", "行政区划码"
]

def main():
    csv_path = sys.argv[1] if len(sys.argv) > 1 else None
    if not csv_path or not os.path.exists(csv_path):
        print(f"Usage: python {sys.argv[0]} <csv_file>")
        sys.exit(1)

    print(f"Reading {csv_path} ...")
    with open(csv_path, "r", encoding="utf-8") as f:
        lines = f.read().strip().splitlines()

    start = 0
    for i, line in enumerate(lines):
        if line.startswith("Location_ID"):
            start = i
            break

    reader = csv.reader(lines[start:])
    rows = list(reader)
    data = rows[1:]
    print(f"  Total cities: {len(data)}")

    wb = Workbook()
    ws = wb.active
    ws.title = "中国城市码"

    hdr_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    hdr_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="D9D9D9")
    cell_border = Border(bottom=thin)
    data_font = Font(name="微软雅黑", size=10)

    for col_idx, title in enumerate(HEADERS_ZH, 1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align

    for row_idx, row in enumerate(data, 2):
        for col_idx, val in enumerate(row, 1):
            if col_idx > len(HEADERS_ZH):
                break
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = cell_border
            if col_idx in (1, 4, 11, 12, 13, 14):
                cell.alignment = Alignment(horizontal="center")

    col_widths = [12, 18, 10, 8, 10, 8, 16, 10, 16, 10, 16, 10, 12, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.auto_filter.ref = f"A1:N{len(data) + 1}"
    ws.freeze_panes = "A2"

    wb.save(OUTPUT)
    print(f"\nSaved to: {OUTPUT}")

if __name__ == "__main__":
    main()
